import openpyxl
import xlsxwriter
import os

# ========== HÀM GHI FILE EXCEL ==========

def LuuExcel(path, dsNV):
    """Ghi danh sách nhân viên vào file Excel"""
    workbook = xlsxwriter.Workbook(path)
    sheet = workbook.add_worksheet("NhanVien")

    # Định dạng tiêu đề
    bold = workbook.add_format({'bold': True, 'bg_color': '#DDEEFF', 'border': 1})

    # Tiêu đề cột
    sheet.write('A1', 'MÃ NV', bold)
    sheet.write('B1', 'TÊN NHÂN VIÊN', bold)
    sheet.write('C1', 'TUỔI', bold)

    # Ghi dữ liệu
    row = 1
    for nv in dsNV:
        sheet.write(row, 0, nv['MaNV'])
        sheet.write(row, 1, nv['TenNV'])
        sheet.write(row, 2, nv['Tuoi'])
        row += 1

    workbook.close()
    print("💾 Đã lưu dữ liệu vào", path)


# ========== HÀM ĐỌC FILE EXCEL ==========

def DocExcel(path):
    """Đọc danh sách nhân viên từ file Excel"""
    if not os.path.exists(path):
        print("⚠️ File chưa tồn tại. Tạo mới khi lưu.")
        return []
    wb = openpyxl.load_workbook(path)
    sheet = wb.active
    ds = []
    for row in sheet.iter_rows(min_row=2, values_only=True):  # bỏ hàng tiêu đề
        if row[0]:
            nv = {"MaNV": row[0], "TenNV": row[1], "Tuoi": int(row[2])}
            ds.append(nv)
    wb.close()
    return ds


# ========== HÀM NGHIỆP VỤ ==========

def XuatDanhSach(dsNV):
    print("\n{:<10}{:<20}{:<10}".format("MÃ NV", "TÊN NHÂN VIÊN", "TUỔI"))
    print("-"*40)
    for nv in dsNV:
        print("{:<10}{:<20}{:<10}".format(nv['MaNV'], nv['TenNV'], nv['Tuoi']))
    print()


def ThemNhanVien(dsNV):
    ma = input("Nhập mã NV: ")
    ten = input("Nhập tên nhân viên: ")
    tuoi = int(input("Nhập tuổi: "))
    dsNV.append({"MaNV": ma, "TenNV": ten, "Tuoi": tuoi})
    print("✅ Đã thêm nhân viên mới.")


def SapXepTheoTuoi(dsNV):
    dsNV.sort(key=lambda x: x['Tuoi'])
    print("✅ Đã sắp xếp theo tuổi tăng dần.")


# ========== CHƯƠNG TRÌNH CHÍNH ==========

def main():
    path = "nhanvien.xlsx"
    dsNV = DocExcel(path)

    while True:
        print("""
========= QUẢN LÝ NHÂN VIÊN =========
1. Xem danh sách
2. Thêm nhân viên
3. Sắp xếp theo tuổi tăng dần
4. Lưu vào file Excel
5. Thoát
====================================
""")
        chon = input("Chọn chức năng (1-5): ")

        if chon == '1':
            XuatDanhSach(dsNV)
        elif chon == '2':
            ThemNhanVien(dsNV)
        elif chon == '3':
            SapXepTheoTuoi(dsNV)
        elif chon == '4':
            LuuExcel(path, dsNV)
        elif chon == '5':
            print("👋 Thoát chương trình.")
            break
        else:
            print("❌ Lựa chọn không hợp lệ!")

if __name__ == "__main__":
    main()
